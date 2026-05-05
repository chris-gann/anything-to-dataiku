"""Excel .xlsx workbook parsing utilities.

Parses workbooks into structured per-sheet data with formula clustering
and light visual-recipe hints so Claude can reason about the flow
without reading every cell.

Design: parser surfaces structure, Claude decides. No planner — Claude
decides consolidation, naming, splitting, and which sheets map to
which recipe type using the RULES in the dataiku MCP server instructions.

Key technique: formula CLUSTERING. A column with 1000 copies of the
same VLOOKUP is returned as ONE pattern with a row range, not 1000
individual formulas. Without this, large workbooks would overwhelm
Claude's context.
"""

import re
import warnings
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ── Formula function → classification ──────────────────────────────────
# Classification determines recipe mapping. Priority order matters:
# a single formula can contain multiple functions; we classify by the
# "heaviest" function found (lookup > aggregation > conditional > ...).

FN_TO_CLASS = {
    # lookup (→ join)
    "VLOOKUP": "lookup", "HLOOKUP": "lookup", "XLOOKUP": "lookup",
    "LOOKUP": "lookup", "MATCH": "lookup", "INDEX": "lookup",
    # conditional aggregation (→ grouping)
    "SUMIF": "conditional_aggregation", "SUMIFS": "conditional_aggregation",
    "COUNTIF": "conditional_aggregation", "COUNTIFS": "conditional_aggregation",
    "AVERAGEIF": "conditional_aggregation", "AVERAGEIFS": "conditional_aggregation",
    "MAXIFS": "conditional_aggregation", "MINIFS": "conditional_aggregation",
    # simple aggregation (→ grouping OR total row, Claude decides)
    "SUM": "aggregation", "AVERAGE": "aggregation", "MIN": "aggregation",
    "MAX": "aggregation", "COUNT": "aggregation", "COUNTA": "aggregation",
    "MEDIAN": "aggregation", "STDEV": "aggregation", "VAR": "aggregation",
    "SUBTOTAL": "aggregation", "AGGREGATE": "aggregation",
    # conditional (→ prepare)
    "IF": "conditional", "IFS": "conditional", "IFERROR": "conditional",
    "IFNA": "conditional", "SWITCH": "conditional",
    "AND": "conditional", "OR": "conditional", "NOT": "conditional",
    # date (→ prepare)
    "DATE": "date", "EDATE": "date", "EOMONTH": "date",
    "YEAR": "date", "MONTH": "date", "DAY": "date",
    "HOUR": "date", "MINUTE": "date", "SECOND": "date",
    "TODAY": "date", "NOW": "date", "WEEKDAY": "date",
    "DATEDIF": "date", "DATEVALUE": "date", "NETWORKDAYS": "date",
    # text (→ prepare)
    "TEXT": "text", "LEFT": "text", "RIGHT": "text", "MID": "text",
    "LEN": "text", "CONCAT": "text", "CONCATENATE": "text", "TEXTJOIN": "text",
    "TRIM": "text", "UPPER": "text", "LOWER": "text", "PROPER": "text",
    "SUBSTITUTE": "text", "REPLACE": "text", "FIND": "text", "SEARCH": "text",
    "VALUE": "text",
    # math (→ prepare)
    "ROUND": "math", "ROUNDUP": "math", "ROUNDDOWN": "math",
    "INT": "math", "ABS": "math", "MOD": "math", "POWER": "math",
    "SQRT": "math", "LOG": "math", "LN": "math", "EXP": "math",
    "RAND": "math", "RANDBETWEEN": "math",
    # array / dynamic (→ python)
    "SUMPRODUCT": "array", "MMULT": "array", "TRANSPOSE": "array",
    "INDIRECT": "dynamic", "OFFSET": "dynamic",
}

# How each classification maps to a Dataiku recipe type.
CLASS_TO_RECIPE = {
    "lookup": "join",
    "conditional_aggregation": "grouping",
    "aggregation": "grouping",  # Claude decides if it's group vs total
    "conditional": "prepare",
    "date": "prepare",
    "text": "prepare",
    "math": "prepare",
    "arithmetic": "prepare",
    "array": "python",
    "dynamic": "python",
    "other": "python",
}

# Priority when classifying a formula with multiple functions.
CLASS_PRIORITY = [
    "array", "dynamic",            # unmappable wins (pushes to python)
    "lookup",
    "conditional_aggregation",
    "aggregation",
    "conditional",
    "date", "text", "math",
    "arithmetic", "other",
]

# Priority when summarizing a whole sheet's mix of formula classes.
SHEET_HINT_PRIORITY = [
    ("array", "python"),
    ("dynamic", "python"),
    ("lookup", "join"),
    ("conditional_aggregation", "grouping"),
    ("aggregation", "grouping"),
    ("conditional", "prepare"),
    ("date", "prepare"),
    ("text", "prepare"),
    ("math", "prepare"),
    ("arithmetic", "prepare"),
]

FN_RE = re.compile(r"\b([A-Z][A-Z0-9\.]{1,})\s*\(")
SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_\. ]*))!")
CELL_ROW_RE = re.compile(r"(\$?[A-Z]+\$?)(\d+)")


# ── Public API ─────────────────────────────────────────────────────────

def parse_workbook(filepath: str) -> dict:
    """Parse an .xlsx file into structured per-sheet data.

    Returns a dict with:
      filename, filepath
      has_macros: bool (True for .xlsm)
      sheets: list of sheet dicts, each with:
          name, dimensions {rows, cols}
          header_row: int or None (1-indexed; heuristic detection)
          headers: list of detected header strings
          sample_rows: first ~5 data rows with COMPUTED values
          cell_stats: {values, formulas, empty}
          formula_patterns: clustered formulas (one entry per
            column×normalized-pattern) with classification, count, and
            range. This is the KEY output for pipeline understanding.
          formula_summary: {class_name: count}
          tables: list of declared Excel Tables {name, range, columns}
          recipe_hint: data_only | prepare | grouping | join | mixed |
            python — best guess at the Dataiku recipe type for this sheet.
      named_ranges: list of {name, value}
      cross_sheet_refs: {sheet_name: [other_sheets_referenced]}
        The dataflow — which sheets feed which.
      source_sheets: sheets with values but no formulas (pure data inputs)
      derived_sheets: sheets with formulas (computed outputs)
    """
    is_macro = filepath.lower().endswith((".xlsm", ".xltm"))
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb_f = load_workbook(filepath, data_only=False)
        wb_v = load_workbook(filepath, data_only=True)

    sheets = [_analyze_sheet(name, wb_f[name], wb_v[name]) for name in wb_f.sheetnames]

    return {
        "filename": Path(filepath).name,
        "filepath": filepath,
        "has_macros": is_macro,
        "sheets": sheets,
        "named_ranges": _extract_named_ranges(wb_f),
        "cross_sheet_refs": _build_cross_sheet_refs(sheets),
        "source_sheets": [s["name"] for s in sheets
                          if s["cell_stats"]["formulas"] == 0
                          and s["cell_stats"]["values"] > 0],
        "derived_sheets": [s["name"] for s in sheets
                           if s["cell_stats"]["formulas"] > 0],
    }


# ── Sheet analysis ─────────────────────────────────────────────────────

def _analyze_sheet(name, ws_f, ws_v) -> dict:
    values_count = formulas_count = empty_count = 0
    # Cluster formulas by (column_letter, normalized_pattern)
    clusters = defaultdict(list)
    raw_formulas_for_class = []

    for row in ws_f.iter_rows():
        for cell in row:
            v = cell.value
            if v is None or (isinstance(v, str) and not v.strip()):
                empty_count += 1
            elif isinstance(v, str) and v.startswith("="):
                formulas_count += 1
                col_letter = cell.column_letter
                pattern = _normalize_formula(v)
                clusters[(col_letter, pattern)].append((cell.row, cell.coordinate))
                raw_formulas_for_class.append(v)
            else:
                values_count += 1

    patterns = []
    class_counter = Counter()
    for (col_letter, pattern), cells in sorted(clusters.items(),
                                               key=lambda kv: (kv[0][0], kv[1][0][0])):
        cls = _classify_formula(pattern)
        rows = [r for r, _ in cells]
        patterns.append({
            "column": col_letter,
            "pattern": pattern,
            "classification": cls,
            "referenced_sheets": sorted(set(_referenced_sheets(pattern))),
            "count": len(cells),
            "row_range": f"{min(rows)}-{max(rows)}" if len(rows) > 1 else str(rows[0]),
            "example_cell": cells[0][1],
        })
        class_counter[cls] += len(cells)

    header_row, headers = _detect_headers(ws_v)
    sample_rows = _sample_data_rows(ws_v, header_row, n=5)

    tables = [
        {"name": t, "range": str(ws_f.tables[t].ref) if hasattr(ws_f.tables[t], "ref") else "",
         "columns": [c.name for c in getattr(ws_f.tables[t], "tableColumns", [])]}
        for t in ws_f.tables
    ]

    return {
        "name": name,
        "dimensions": {"rows": ws_f.max_row, "cols": ws_f.max_column},
        "header_row": header_row,
        "headers": headers,
        "sample_rows": sample_rows,
        "cell_stats": {
            "values": values_count,
            "formulas": formulas_count,
            "empty": empty_count,
        },
        "formula_patterns": patterns,
        "formula_summary": dict(class_counter),
        "tables": tables,
        "recipe_hint": _sheet_recipe_hint(values_count, formulas_count, class_counter),
    }


# ── Formula helpers ────────────────────────────────────────────────────

def _normalize_formula(f: str) -> str:
    """Replace row numbers in cell references with N to enable clustering.

    =VLOOKUP(A17, Sheet2!A:C, 2, FALSE)  →  =VLOOKUP(AN, Sheet2!A:C, N, FALSE)
    Note: also normalizes bare integer literals (e.g. the 2 above). That's
    intentional — formulas that differ only in row numbers or small int
    literals are almost always the same logical operation.
    """
    return CELL_ROW_RE.sub(lambda m: f"{m.group(1)}N", f)


def _classify_formula(f: str) -> str:
    if f.startswith("{=") or f.endswith("}"):
        return "array"
    fns = [fn.upper() for fn in FN_RE.findall(f)]
    classes = {FN_TO_CLASS.get(fn, "other") for fn in fns}
    for candidate in CLASS_PRIORITY:
        if candidate in classes:
            return candidate
    if not fns:
        return "arithmetic"
    return "other"


def _referenced_sheets(f: str) -> list:
    out = []
    for quoted, bare in SHEET_REF_RE.findall(f):
        name = quoted or bare
        if name and not name.upper().startswith(("TRUE", "FALSE")):
            out.append(name.strip())
    return out


# ── Sheet-level heuristics ─────────────────────────────────────────────

def _sheet_recipe_hint(values: int, formulas: int, cls: Counter) -> str:
    if formulas == 0:
        return "data_only" if values > 0 else "empty"
    for class_name, recipe in SHEET_HINT_PRIORITY:
        if cls.get(class_name, 0) > 0:
            # If multiple recipe types are represented, call it mixed.
            recipes_present = {CLASS_TO_RECIPE[c] for c in cls if c in CLASS_TO_RECIPE}
            if len(recipes_present - {"prepare"}) > 1:
                return "mixed"
            return recipe
    return "python"


def _detect_headers(ws_v):
    """Return (header_row_1indexed, header_strings).

    Heuristic: within the first 10 rows, find the row with the most
    non-empty string cells AND non-empty data below it. If nothing is
    confidently a header, return (None, []).
    """
    best = (None, [], -1)
    max_scan = min(ws_v.max_row, 10)
    for r in range(1, max_scan + 1):
        row_vals = [ws_v.cell(r, c).value for c in range(1, ws_v.max_column + 1)]
        string_count = sum(1 for v in row_vals if isinstance(v, str) and v.strip())
        if string_count < 2:
            continue
        # Check that the next row has data
        next_r = r + 1
        if next_r > ws_v.max_row:
            continue
        next_vals = [ws_v.cell(next_r, c).value for c in range(1, ws_v.max_column + 1)]
        if not any(v is not None for v in next_vals):
            continue
        score = string_count
        if score > best[2]:
            headers = [str(v).strip() if v is not None else "" for v in row_vals]
            best = (r, headers, score)
    return best[0], best[1]


def _sample_data_rows(ws_v, header_row, n=5):
    start = (header_row + 1) if header_row else 1
    out = []
    for r in range(start, min(ws_v.max_row + 1, start + n)):
        row = [ws_v.cell(r, c).value for c in range(1, ws_v.max_column + 1)]
        if any(v is not None for v in row):
            out.append(row)
    return out


# ── Workbook-level helpers ─────────────────────────────────────────────

def _extract_named_ranges(wb) -> list:
    out = []
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        out.append({"name": name, "value": dn.value if hasattr(dn, "value") else str(dn)})
    return out


def _build_cross_sheet_refs(sheets: list) -> dict:
    """For each sheet, which other sheets does it reference in formulas?"""
    out = {}
    sheet_names = {s["name"] for s in sheets}
    for s in sheets:
        refs = set()
        for p in s["formula_patterns"]:
            for ref in p["referenced_sheets"]:
                if ref in sheet_names and ref != s["name"]:
                    refs.add(ref)
        if refs:
            out[s["name"]] = sorted(refs)
    return out
